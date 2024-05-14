from openpyxl import load_workbook, Workbook

def fill_excel_with_data(excel_path, extracted_data):
    print(f"Filling Excel with data: {extracted_data}")
    amount, distance, time, date = extracted_data

    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 查找里程、打车日期、打车时间、每笔实际打车费对应的列
    distance_col = None
    date_col = None
    time_col = None
    actual_fare_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == '行程距离':
            distance_col = col
        elif header == '打车日期':
            date_col = col
        elif header == '打车时间':
            time_col = col
        elif header == '每笔实际打车费':
            actual_fare_col = col

    if not all([distance_col, date_col, time_col, actual_fare_col]):
        print("所需列未找到")
        return

    # 若未找到匹配的金额行，则写入新的工作表
    if '未匹配数据' not in workbook.sheetnames:
        new_sheet = workbook.create_sheet('未匹配数据')
        # 添加表头
        new_sheet.append(['行程距离', '打车日期', '打车时间', '每笔实际打车费'])
    else:
        new_sheet = workbook['未匹配数据']

    # 获取新工作表的当前最大行数
    start_row = new_sheet.max_row + 1
    if start_row > 1:
        # 在每个数据块之间插入4行空行
        start_row += 4

    # 填入PDF表格数据
    new_sheet.cell(row=start_row, column=1).value = distance
    new_sheet.cell(row=start_row, column=2).value = date
    new_sheet.cell(row=start_row, column=3).value = time
    new_sheet.cell(row=start_row, column=4).value = amount

    print(f"Data written to Excel at row {start_row}")

    # 保存修改后的Excel文件
    workbook.save(excel_path)
    print(f"Excel file saved: {excel_path}")
