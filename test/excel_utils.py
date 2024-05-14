from openpyxl import load_workbook, Workbook

def fill_excel_with_data(excel_path, extracted_data, total_amount):
    # 打开Excel文件
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 查找核定报销金额列
    reimbursement_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == '核定报销金额':
            reimbursement_col = col
            break

    if reimbursement_col is None:
        print("核定报销金额列未找到")
        return

    # 查找里程、打车日期、起点+终点、每笔实际打车费对应的列
    distance_col = None
    date_col = None
    time_col = None
    service_content_col = None
    actual_fare_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == '行程距离':
            distance_col = col
        elif header == '打车日期':
            date_col = col
        elif header == '打车时间':
            time_col = col
        elif header == '服务内容':
            service_content_col = col
        elif header == '每笔实际打车费':
            actual_fare_col = col

    if not all([distance_col, date_col, time_col, service_content_col, actual_fare_col]):
        print("所需列未找到")
        return

    # 根据金额判断填入行
    target_row = None
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=reimbursement_col).value == total_amount:
            target_row = r
            break

    if target_row is None:
        # 若未找到匹配的金额行，则写入新的工作表
        if '未匹配数据' not in workbook.sheetnames:
            new_sheet = workbook.create_sheet('未匹配数据')
            # 添加表头
            new_sheet.append(['服务内容', '打车日期', '打车时间', '每笔实际打车费', '行程距离'])
        else:
            new_sheet = workbook['未匹配数据']

        # 获取新工作表的当前最大行数
        start_row = new_sheet.max_row + 1
        if start_row > 1:
            # 在每个数据块之间插入4行空行
            start_row += 4

        # 填入PDF表格数据
        for i, row in enumerate(extracted_data.iterrows(), start=start_row - 1):
            date_time_str = row[1]['上车时间']
            date_part = date_time_str.split()[0]
            time_part = date_time_str.split()[1]

            new_sheet.cell(row=i, column=1).value = f"{row[1]['起点']} - {row[1]['终点']}"
            new_sheet.cell(row=i, column=2).value = date_part
            new_sheet.cell(row=i, column=3).value = time_part
            new_sheet.cell(row=i, column=4).value = row[1]['金额[元]']
            new_sheet.cell(row=i, column=5).value = row[1]['里程[公里]']
    else:
        # 确定插入数据的起始行和插入行数
        start_row = target_row + 1
        num_rows = len(extracted_data)

        # 在target_row处插入num_rows行的空行
        sheet.insert_rows(start_row, num_rows)

        # 填入PDF表格数据
        for i, row in enumerate(extracted_data.iterrows(), start=start_row):
            date_time_str = row[1]['上车时间']
            date_part = date_time_str.split()[0]
            time_part = date_time_str.split()[1]

            sheet.cell(row=i, column=distance_col).value = row[1]['里程[公里]']
            sheet.cell(row=i, column=date_col).value = date_part
            sheet.cell(row=i, column=time_col).value = time_part
            sheet.cell(row=i, column=service_content_col).value = f"{row[1]['起点']} - {row[1]['终点']}"
            sheet.cell(row=i, column=actual_fare_col).value = row[1]['金额[元]']

    # 保存修改后的Excel文件
    workbook.save(excel_path)
