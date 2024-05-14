import pdfplumber
import pandas as pd
from openpyxl import load_workbook


def extract_tables_from_pdf(pdf_path):
    all_data = []
    total_amount = 0

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            # 提取“共x笔行程，合计xxx元”的信息
            for line in text.split('\n'):
                if '合计' in line:
                    total_amount += float(line.split('合计')[-1].replace('元', '').strip())
                    break

            table = page.extract_table()

            # 打印提取的表格以便调试
            print("Extracted table from page:", table)

            if table:
                headers = table[0]  # 表头
                data = table[1:]  # 数据
                df = pd.DataFrame(data, columns=headers)
                all_data.append(df)

    if all_data:
        combined_df = pd.concat(all_data, ignore_index=True)
    else:
        combined_df = pd.DataFrame()

    return combined_df, total_amount


def fill_excel_with_data(excel_path, extracted_data, total_amount):
    # 打开Excel文件
    workbook = load_workbook(excel_path)
    sheet = workbook.active

    # 查找核定报销金额列
    reimbursement_col = None
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == '核定报销金额':
            reimbursement_col = col
            break

    if reimbursement_col is None:
        print("核定报销金额列未找到")
        return

    # 查找里程、打车日期、起点+终点、每笔实际打车费对应的列
    distance_col = None
    date_col = None
    service_content_col = None
    actual_fare_col = None
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header == '行程距离':
            distance_col = col
        elif header == '打车日期':
            date_col = col
        elif header == '打车时间':
            time_col = col
        elif header == '服务内容':
            service_content_col = col
        elif header == '每笔实际打车费':
            actual_fare_col = col

    if not all([distance_col, date_col, service_content_col, actual_fare_col]):
        print("所需列未找到")
        return

    # 根据金额判断填入行
    target_row = None
    for r in range(2, sheet.max_row + 1):
        if sheet.cell(row=r, column=reimbursement_col).value == total_amount:
            target_row = r
            break

    if target_row is None:
        print("未找到匹配的金额行")
        return

    # 确定插入数据的起始行和插入行数
    start_row = target_row + 1
    num_rows = len(extracted_data)

    # 在target_row处插入num_rows行的空行
    sheet.insert_rows(start_row, num_rows - 1)

    # 填入PDF表格数据
    for i, row in enumerate(extracted_data.iterrows(), start=start_row - 1):
        # 分割日期和时间
        date_time_str = row[1]['上车时间']
        date_part = date_time_str.split()[0]
        time_part = date_time_str.split()[1]

        sheet.cell(row=i, column=distance_col).value = row[1]['里程[公里]']
        sheet.cell(row=i, column=date_col).value = date_part
        sheet.cell(row=i, column=time_col).value = time_part
        sheet.cell(row=i, column=service_content_col).value = f"{row[1]['起点']} - {row[1]['终点']}"
        sheet.cell(row=i, column=actual_fare_col).value = row[1]['金额[元]']

    # 保存修改后的Excel文件
    workbook.save(excel_path)


# 使用函数读取PDF并提取表格和合计金额
pdf_path = '2.pdf'
table_df, total_amount = extract_tables_from_pdf(pdf_path)

# 输出提取的表格
print(table_df)

# 提取所需字段并填入Excel
if not table_df.empty and total_amount is not None:
    excel_path = 'target.xlsx'
    fill_excel_with_data(excel_path, table_df, total_amount)